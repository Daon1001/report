"""
CRETOP PDF 파서: 기업 브리핑(개요) 및 신용등급 보고서에서 데이터 추출
"""
import re
import os
from typing import Dict, Any, Optional

try:
    import pdfplumber
except ImportError:
    pdfplumber = None


def parse_company_overview(filepath: str) -> Dict[str, Any]:
    """개요.pdf (기업 브리핑 보고서) 파싱 - 다양한 기업에 대응"""
    if pdfplumber is None or filepath is None:
        return _fallback_overview()
    
    data = {
        "기업명": "", "사업자번호": "", "대표자명": "", "법인번호": "",
        "설립일자": "", "종업원수": "", "주소": "", "전화번호": "",
        "기업유형": "", "표준산업분류": "", "기업규모": "", "주요제품": "",
        "기업신용등급": "", "EW등급": "",
        "신용정보": {}, "주요주주": [], "주요구매처": [], "주요판매처": [],
        "기업인증": {}, "산업재산권": {}, "재무상태표": {}, "손익계산서": {},
        "재무비율": {}, "재무진단": {},
    }
    
    try:
        with pdfplumber.open(filepath) as pdf:
            full_text = ""
            for page in pdf.pages:
                text = page.extract_text() or ""
                full_text += text + "\n"
            
            if not full_text.strip():
                return data
            
            # ── 기업 기본정보 ──
            patterns = {
                "사업자번호": r"사업자번호\s+([\d\-]+)",
                "법인번호": r"법인\(주민\)번호\s+([\d\-]+)",
                "대표자명": r"대표자명\s+(\S+)",
                "설립일자": r"설립일자\s+([\d\-]+)",
                "종업원수": r"종업원수\s+(\d+\s*명)",
                "전화번호": r"전화번호\s+([\d\-]+)",
                "기업규모": r"기업규모\s+(\S+)",
            }
            for key, pattern in patterns.items():
                match = re.search(pattern, full_text)
                if match:
                    data[key] = match.group(1).strip()
            
            # 기업명 - (주)XXX 또는 주식회사XXX 패턴
            name_match = re.search(r"\(주\)\S+", full_text)
            if name_match:
                data["기업명"] = name_match.group(0)
            else:
                name_match2 = re.search(r"기업명\s*[:：]?\s*(\S+)", full_text)
                if name_match2:
                    data["기업명"] = name_match2.group(1)
            
            # 기업유형
            type_match = re.search(r"기업유형/형태\s+(.+?)(?:\s+전화번호|\n)", full_text)
            if type_match:
                data["기업유형"] = type_match.group(1).strip()
            
            # 주소 - 더 범용적 패턴
            addr_match = re.search(r"주소\(도로명\)\s+(.+?)(?:\s+주요제품|\s+전화번호|\n)", full_text)
            if not addr_match:
                addr_match = re.search(r"주소\(도로명\)\s+(\(.+?\).*?)(?:\n)", full_text)
            if addr_match:
                data["주소"] = addr_match.group(1).strip()
            
            # 표준산업분류
            ind_match = re.search(r"표준산업분류\(11차\)\s+(\(.+?\).+?)(?:\s+주요제품|\n)", full_text)
            if ind_match:
                data["표준산업분류"] = ind_match.group(1).strip()
            
            # 주요제품
            prod_match = re.search(r"주요제품\(상품\)\s+(.+?)(?:\n)", full_text)
            if prod_match:
                data["주요제품"] = prod_match.group(1).strip()
            
            # 기업신용등급 - 소문자/대문자 모두 (a, bb-, BBB+ 등)
            grade_match = re.search(r"평가일자\s*:\s*([\d\-]+)\s*\n?\s*결산일자\s*:\s*([\d\-]+)", full_text)
            if grade_match:
                data["신용등급_평가일자"] = grade_match.group(1)
                data["신용등급_결산일자"] = grade_match.group(2)
            
            # 신용등급 값 추출 (게이지 옆의 등급 텍스트)
            credit_grade = re.search(r"\n\s*([a-dA-D]{1,3}[+-]?)\s*\n.*?평가일자", full_text, re.DOTALL)
            if credit_grade:
                data["기업신용등급"] = credit_grade.group(1).strip()
            else:
                # bb-, a, bbb+ 등 패턴
                credit_grade2 = re.search(r"기업신용등급.*?\n\s*([a-dA-D]{1,3}[+-]?)\s*\n", full_text, re.DOTALL)
                if credit_grade2:
                    data["기업신용등급"] = credit_grade2.group(1).strip()
            
            # EW등급 - 정상/주의/경고/위험/유보 모두 인식
            ew_match = re.search(r"(정상|주의|경고|위험|유보)\s*[▼▲►▶]?\s*\n.*?기준일자", full_text, re.DOTALL)
            if ew_match:
                data["EW등급"] = ew_match.group(1)
            else:
                ew_match2 = re.search(r"EW\s*등급.*?(정상|주의|경고|위험|유보)", full_text, re.DOTALL)
                if ew_match2:
                    data["EW등급"] = ew_match2.group(1)
                else:
                    # 단순 검색
                    for ew_val in ["정상", "유보", "주의", "경고", "위험"]:
                        if ew_val in full_text:
                            data["EW등급"] = ew_val
                            break
            
            # 재무진단 등급 - "보통 이하" 등도 인식
            diag_patterns = {
                "성장성": r"성장성\s+(우수|양호|보통[^\n]*|미흡|열위)",
                "수익성": r"수익성\s+(우수|양호|보통[^\n]*|미흡|열위)",
                "재무구조": r"재무구조\s+(우수|양호|보통[^\n]*|미흡|열위)",
                "부채상환능력": r"부채상환능력\s+(우수|양호|보통[^\n]*|미흡|열위|등급없음)",
                "활동성": r"활동성\s+(우수|양호|보통[^\n]*|미흡|열위)",
            }
            for key, pattern in diag_patterns.items():
                match = re.search(pattern, full_text)
                if match:
                    val = match.group(1).strip()
                    # "보통 이하" → "보통"으로 정규화
                    if "보통" in val:
                        val = "보통"
                    data["재무진단"][key] = val
    
    except Exception as e:
        print(f"PDF 파싱 오류: {e}")
    
    return data


def parse_credit_report(filepath: str) -> Dict[str, Any]:
    """신용.pdf (기업 신용등급 보고서) 파싱"""
    data = {
        "현재등급": "",
        "등급설명": "",
        "등급구분": "",
        "평가일자": "",
        "재무기준일자": "",
        "등급이력": [],
        "현금흐름등급": [],
        "외부신용등급": {},
    }
    
    if pdfplumber is None:
        return data
    
    try:
        with pdfplumber.open(filepath) as pdf:
            full_text = ""
            for page in pdf.pages:
                text = page.extract_text() or ""
                full_text += text + "\n"
            
            # 현재 신용등급
            grade_match = re.search(r"등급\s+평가\(산출\)일자\s+재무기준일자\s+등급구분\s*\n\s*(\S+)\s+([\d\-]+)\s+([\d\-]+)\s+(\S+)", full_text)
            if grade_match:
                data["현재등급"] = grade_match.group(1)
                data["평가일자"] = grade_match.group(2)
                data["재무기준일자"] = grade_match.group(3)
                data["등급구분"] = grade_match.group(4)
            
            # 등급설명
            desc_match = re.search(r"등급설명\s+(.+?)(?:\n등급구분)", full_text, re.DOTALL)
            if desc_match:
                data["등급설명"] = desc_match.group(1).strip()
            
            # 현금흐름등급 - 연도별
            cr_pattern = r"(20\d{2}-\d{2}-\d{2})\s+(?:.*?)(CR\d|V|판정보류|판정제외)"
            for match in re.finditer(cr_pattern, full_text):
                data["현금흐름등급"].append({
                    "기준일자": match.group(1),
                    "등급": match.group(2)
                })
    except Exception as e:
        print(f"신용 PDF 파싱 오류: {e}")
    
    return data


def _fallback_overview() -> Dict[str, Any]:
    """기본 빈 값 반환"""
    return {
        "기업명": "", "사업자번호": "", "대표자명": "", "법인번호": "",
        "설립일자": "", "종업원수": "", "주소": "", "전화번호": "",
        "기업유형": "", "표준산업분류": "", "기업규모": "", "주요제품": "",
        "기업신용등급": "", "EW등급": "", "재무진단": {},
    }


def parse_company_overview_excel(filepath: str) -> Dict[str, Any]:
    """기업개요 엑셀(report.xls) 파싱 — CRETOP 엑셀 다운로드 형식"""
    import pandas as pd
    import numpy as np
    
    data = _fallback_overview()
    
    try:
        # xls → libreoffice 변환이 필요할 수 있음. 일단 pandas로 시도
        try:
            xls = pd.ExcelFile(filepath)
        except ImportError:
            # xlrd가 없으면 libreoffice로 변환
            import subprocess, tempfile, os
            tmp_xlsx = tempfile.mktemp(suffix='.xlsx')
            tmp_dir = os.path.dirname(tmp_xlsx)
            subprocess.run(['libreoffice', '--headless', '--convert-to', 'xlsx', 
                          '--outdir', tmp_dir, filepath], 
                         capture_output=True, timeout=30)
            converted = filepath.rsplit('.', 1)[0] + '.xlsx'
            if not os.path.exists(converted):
                # 파일명이 다를 수 있음
                import glob
                candidates = glob.glob(os.path.join(tmp_dir, '*.xlsx'))
                converted = candidates[-1] if candidates else filepath
            xls = pd.ExcelFile(converted)
        
        sheets = xls.sheet_names
        
        # 모든 시트의 텍스트를 합쳐서 검색
        all_cells = []
        for sheet in sheets:
            df = pd.read_excel(xls, sheet_name=sheet, header=None)
            for r in range(df.shape[0]):
                for c in range(df.shape[1]):
                    val = df.iloc[r, c]
                    if pd.notna(val):
                        all_cells.append((str(val).strip(), r, c, sheet, df))
        
        # 키-값 쌍 추출 (키 셀 옆이나 같은 행에서 값 찾기)
        def find_value_after(keyword):
            for val, r, c, sheet, df in all_cells:
                # 키워드가 셀에 포함되어 있으면
                if keyword in val:
                    # 같은 셀에 값도 있는 경우 ("사업자번호 215-88-00406")
                    cleaned = val.replace(keyword, '').strip(' :：-·・').strip()
                    if cleaned and cleaned != val and len(cleaned) > 1:
                        return cleaned
                    # 같은 행의 오른쪽 셀들 탐색 (최대 15칸)
                    for cc in range(c+1, min(c+15, df.shape[1])):
                        rv = df.iloc[r, cc]
                        if pd.notna(rv):
                            sv = str(rv).strip()
                            if sv and sv != keyword and len(sv) > 0:
                                return sv
            return ""
        
        # (주)XXX 패턴으로 기업명 직접 찾기
        company_name = ""
        for val, r, c, s, d in all_cells:
            if val.startswith("(주)") and len(val) > 3:
                company_name = val
                break
        
        data["기업명"] = company_name or find_value_after("기업명")
        
        # 사업자번호 - "215-88-00406" 패턴 직접 찾기
        import re as _re
        bizno = find_value_after("사업자번호")
        if not bizno or len(bizno) < 10:
            for val, r, c, s, d in all_cells:
                if _re.match(r'^\d{3}-\d{2}-\d{5}$', val):
                    bizno = val
                    break
        data["사업자번호"] = bizno
        data["대표자명"] = find_value_after("대표자명")
        data["법인번호"] = find_value_after("법인(주민)번호")
        data["설립일자"] = find_value_after("설립일자")
        data["종업원수"] = find_value_after("종업원수")
        data["주소"] = find_value_after("주소(도로명)")
        data["전화번호"] = find_value_after("전화번호")
        data["기업유형"] = find_value_after("기업유형/형태")
        data["표준산업분류"] = find_value_after("표준산업분류")
        data["기업규모"] = find_value_after("기업규모")
        data["주요제품"] = find_value_after("주요제품(상품)")
        
        # 재무진단
        for diag in ["성장성", "수익성", "재무구조", "부채상환능력", "활동성"]:
            val = find_value_after(diag)
            if val and val in ["우수", "양호", "보통", "미흡", "열위", "보통 이하", "등급없음"]:
                if "보통" in val:
                    val = "보통"
                data["재무진단"][diag] = val
        
        # 신용등급 (bb-, a, BBB+ 등) - 텍스트 먼저 시도
        for val, r, c, s, d in all_cells:
            if re.match(r'^[a-dA-D]{1,3}[+-]?$', val) and len(val) <= 4:
                data["기업신용등급"] = val
                break
        
        # EW등급 - '법인등기정보 정상'과 구분
        for val, r, c, s, d in all_cells:
            if val in ("정상", "유보", "주의", "경고", "위험"):
                row_texts = set()
                for cc in range(d.shape[1]):
                    rv = d.iloc[r, cc]
                    if pd.notna(rv):
                        row_texts.add(str(rv).strip())
                if '법인등기정보' in ' '.join(row_texts):
                    continue
                data["EW등급"] = val
                break
        
        # ── 이미지 추출 (신용등급 게이지, EW등급, 재무진단) ──
        import base64
        try:
            from openpyxl import load_workbook as _load_wb
            import subprocess, tempfile, glob
            
            _wb = None
            
            # 전략: 항상 xlsx로 변환 후 openpyxl로 읽기
            # (원본이 .xls든 .xlsx든 libreoffice로 변환하면 확실)
            tmp_dir = tempfile.mkdtemp()
            convert_src = filepath  # 원본 파일 경로
            
            # pandas ExcelFile에서 원본 경로 가져오기
            if hasattr(xls, 'io') and isinstance(xls.io, str) and os.path.exists(xls.io):
                convert_src = xls.io
            
            print(f"[이미지추출] 원본: {convert_src}")
            
            # 먼저 openpyxl로 직접 시도 (.xlsx인 경우)
            try:
                _wb = _load_wb(convert_src)
                print(f"[이미지추출] openpyxl 직접 로드 성공")
            except Exception as e1:
                print(f"[이미지추출] openpyxl 직접 실패: {e1}")
                # libreoffice로 변환
                try:
                    proc = subprocess.run(
                        ['libreoffice', '--headless', '--convert-to', 'xlsx', '--outdir', tmp_dir, convert_src],
                        capture_output=True, timeout=60, text=True
                    )
                    print(f"[이미지추출] libreoffice 변환: rc={proc.returncode}, stderr={proc.stderr[:200]}")
                    converted_files = glob.glob(os.path.join(tmp_dir, '*.xlsx'))
                    print(f"[이미지추출] 변환된 파일: {converted_files}")
                    if converted_files:
                        _wb = _load_wb(converted_files[0])
                        print(f"[이미지추출] 변환 후 로드 성공")
                except Exception as e2:
                    print(f"[이미지추출] libreoffice 변환 실패: {e2}")
            
            if _wb:
                # 모든 시트에서 이미지 수집
                sheet_images = {}
                for si, sname in enumerate(_wb.sheetnames):
                    _ws = _wb[sname]
                    for ii, img in enumerate(_ws._images):
                        try:
                            img_bytes = img._data()
                            sz = len(img_bytes)
                            sheet_images.setdefault(si, []).append((ii, img_bytes, sz))
                        except:
                            pass
                
                total_imgs = sum(len(v) for v in sheet_images.values())
                print(f"[이미지추출] 총 {total_imgs}개 이미지 추출")
                
                # 시트 1 (index 1): 기업프로필 → 신용등급 + EW등급
                if 1 in sheet_images:
                    imgs = sheet_images[1]
                    # 로고(45KB급)와 배경(600KB+) 제외
                    small_imgs = [(ii, ib, sz) for ii, ib, sz in imgs if 3000 < sz < 20000]
                    small_imgs.sort(key=lambda x: x[0])
                    print(f"[이미지추출] 시트1 소형 이미지: {len(small_imgs)}개, 크기: {[s[2] for s in small_imgs]}")
                    if len(small_imgs) >= 2:
                        data["신용등급_이미지"] = base64.b64encode(small_imgs[0][1]).decode('utf-8')
                        data["EW등급_이미지"] = base64.b64encode(small_imgs[1][1]).decode('utf-8')
                    elif len(small_imgs) == 1:
                        data["신용등급_이미지"] = base64.b64encode(small_imgs[0][1]).decode('utf-8')
                
                # 마지막 시트: 재무진단 (도넛 차트 5개)
                last_si = len(_wb.sheetnames) - 1
                if last_si in sheet_images and last_si > 1:
                    imgs = sheet_images[last_si]
                    donut_imgs = [(ii, ib, sz) for ii, ib, sz in imgs if 3000 < sz < 15000]
                    donut_imgs.sort(key=lambda x: x[0])
                    print(f"[이미지추출] 마지막시트 도넛: {len(donut_imgs)}개")
                    diag_keys = ["성장성", "수익성", "재무구조", "부채상환능력", "활동성"]
                    for idx, dk in enumerate(diag_keys):
                        if idx < len(donut_imgs):
                            data.setdefault("재무진단_이미지", {})[dk] = base64.b64encode(donut_imgs[idx][1]).decode('utf-8')
                
                print(f"[이미지추출] 결과: 신용등급={'있음' if data.get('신용등급_이미지') else '없음'}, EW={'있음' if data.get('EW등급_이미지') else '없음'}, 재무진단={len(data.get('재무진단_이미지', {}))}개")
            else:
                print("[이미지추출] _wb가 None - 이미지 추출 불가")
                
        except Exception as e:
            import traceback
            print(f"[이미지추출] 최종 에러: {e}")
            traceback.print_exc()
    
    except Exception as e:
        print(f"기업개요 엑셀 파싱 오류: {e}")
    
    return data


def parse_credit_report_excel(filepath: str) -> Dict[str, Any]:
    """신용등급 엑셀(report__1_.xls) 파싱"""
    import pandas as pd
    import numpy as np
    
    data = {
        "현재등급": "", "등급설명": "", "등급구분": "",
        "평가일자": "", "재무기준일자": "",
        "등급이력": [], "현금흐름등급": [], "외부신용등급": {},
    }
    
    try:
        try:
            xls = pd.ExcelFile(filepath)
        except ImportError:
            import subprocess, tempfile, os
            tmp_dir = tempfile.mkdtemp()
            subprocess.run(['libreoffice', '--headless', '--convert-to', 'xlsx',
                          '--outdir', tmp_dir, filepath],
                         capture_output=True, timeout=30)
            import glob
            candidates = glob.glob(os.path.join(tmp_dir, '*.xlsx'))
            xls = pd.ExcelFile(candidates[0]) if candidates else None
            if not xls:
                return data
        
        all_cells = []
        for sheet in xls.sheet_names:
            df = pd.read_excel(xls, sheet_name=sheet, header=None)
            for r in range(df.shape[0]):
                for c in range(df.shape[1]):
                    val = df.iloc[r, c]
                    if pd.notna(val):
                        all_cells.append((str(val).strip(), r, c, sheet, df))
        
        # 등급 이력 (bb-, a 등 + 날짜 패턴)
        for val, r, c, sheet, df in all_cells:
            if re.match(r'^[a-dA-D]{1,3}[+-]?$', val) and len(val) <= 4:
                if not data["현재등급"]:
                    data["현재등급"] = val
                # 같은 행에서 날짜 찾기
                dates = []
                for cc in range(df.shape[1]):
                    cv = df.iloc[r, cc]
                    if pd.notna(cv) and re.match(r'20\d{2}-\d{2}-\d{2}', str(cv).strip()):
                        dates.append(str(cv).strip())
                if dates:
                    data["등급이력"].append({
                        "등급": val,
                        "평가일자": dates[0] if dates else "",
                        "재무기준일자": dates[1] if len(dates) > 1 else "",
                    })
        
        # 등급설명
        for val, r, c, s, d in all_cells:
            if "채무상환능력" in val:
                data["등급설명"] = val
                break
        
        # 평가일자
        for val, r, c, s, d in all_cells:
            if "평가일자" in val and ":" in val:
                date_match = re.search(r'(\d{4}-\d{2}-\d{2})', val)
                if date_match:
                    data["평가일자"] = date_match.group(1)
            elif "결산일자" in val and ":" in val:
                date_match = re.search(r'(\d{4}-\d{2}-\d{2})', val)
                if date_match:
                    data["재무기준일자"] = date_match.group(1)
        
        # 현금흐름등급
        for val, r, c, sheet, df in all_cells:
            if re.match(r'^CR\d$', val):
                # 같은 행에서 날짜 찾기
                for cc in range(df.shape[1]):
                    cv = df.iloc[r, cc]
                    if pd.notna(cv) and re.match(r'20\d{2}-\d{2}-\d{2}', str(cv).strip()):
                        data["현금흐름등급"].append({
                            "기준일자": str(cv).strip(),
                            "등급": val
                        })
                        break
    
    except Exception as e:
        print(f"신용등급 엑셀 파싱 오류: {e}")
    
    return data
